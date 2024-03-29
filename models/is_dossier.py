# -*- coding: utf-8 -*-
from odoo import api, fields, models, _
from datetime import date, datetime
from odoo.exceptions import Warning
import base64
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, PatternFill, Font, borders
from openpyxl.styles.borders import Border
from openpyxl.worksheet.page import PageMargins




def _get_prefix_annee():
    annee = date.today().year
    annee = annee-1989
    x = str(date.today())[-5:]
    if x>='04-01':
        annee+=1 # Si 1er avril, prendre l'année suivante
    return(str(annee))



class IsDossier(models.Model):
    _name = 'is.dossier'
    _description = "Dossier"
    _order = 'numaff desc'
    _rec_name = 'numaff'


    @api.multi
    def _get_numaff(self):
        prefix = _get_prefix_annee()+'A'
        dossiers=self.env['is.dossier'].search([('numaff','like',prefix)],order="numaff desc",limit=1)
        numaff=1
        for dossier in dossiers:
            numaff = int(dossier.numaff[-3:])
            numaff+=1
        numaff=prefix+('000'+str(numaff))[-3:]
        return numaff

    numaff               = fields.Char("Dossier", required=True, index=True, default=lambda self: self._get_numaff())
    nom                  = fields.Char("Nom", required=True)
    description_succinte = fields.Text("Description succinte")
    adresse1             = fields.Char("Adresse 1")
    adresse2             = fields.Char("Adresse 2")
    codepostal           = fields.Char("Code postal")
    lieu                 = fields.Char("Ville")
    numconcours          = fields.Char("Numéro de concours")
    m_ouvrage_id         = fields.Many2one('res.partner', u"Maitre d'ouvrage", domain=[('is_company','=',True)])
    m_oeuvre_id          = fields.Many2one('res.partner', u"Maitre d'oeuvre" , domain=[('is_company','=',True)])
    date                 = fields.Date("Date")
    surface              = fields.Float("Surface")
    montant_ope          = fields.Float("Montant de l'opération")
    site_internet        = fields.Char("Site internet")
    description_complete = fields.Text("Description complète")
    nota                 = fields.Text("Note")
    dossierreference     = fields.Selection([('oui','Oui'),('non','Non')],"Dossier référénce")
    fiche                = fields.Selection([('oui','Oui'),('non','Non')],"Fiche")
    distance             = fields.Float("Distance")
    duree                = fields.Float("Durée")
    service_executant    = fields.Char(u"Service exécutant", help=u"Nécessaire pour Chorus Pro")
    num_marche           = fields.Char(u"N° marché"        , help=u"Nécessaire pour Chorus Pro")
    note_ids             = fields.One2many('is.dossier.note', 'dossier_id', u'Notes')
    referencee_ids       = fields.Many2many('is.dossier.reference', 'is_dossier_is_reference_rel', 'dossier_id', 'reference_id', u'Références')
    contrat_ids          = fields.One2many('is.dossier.contrat', 'dossier_id', u'Contrats')
    adjudicataire_ids    = fields.One2many('is.dossier.adjudicataire', 'dossier_id', u'Adjudicataires')
    state                = fields.Selection([
            ('en cours' ,'En cours'),
            ('termine'  ,'Terminé'),
            ('abandonne','Abandonné'),
        ],"Etat", default="en cours")

    @api.multi
    def name_get(self):
        result = []
        for obj in self:
            result.append((obj.id, str(obj.numaff)+u' - '+str(obj.nom)))
        return result


    @api.model
    def _name_search(self, name, args=None, operator='ilike', limit=100, name_get_uid=None):
        args = args or []
        ids = []
        if name:
            ids = self._search(['|',('numaff', 'ilike', name),('nom', 'ilike', name)] + args, limit=limit, access_rights_uid=name_get_uid)
        else:
            ids = self._search(args, limit=limit, access_rights_uid=name_get_uid)
        return self.browse(ids).name_get()


    @api.multi
    def creation_contrat_action(self, vals):
        for obj in self:
            res= {
                'name': 'Contrat',
                'view_mode': 'form',
                'view_type': 'form',
                'res_model': 'is.dossier.contrat',
                'type': 'ir.actions.act_window',
                'context': {'default_dossier_id': obj.id }
            }
            return res


class IsDossierAdjudicataire(models.Model):
    _name = 'is.dossier.adjudicataire'
    _description = "Dossier Adjudicataire"
    _order = 'id desc'

    dossier_id       = fields.Many2one('is.dossier', 'Dossier', required=True)
    retenu           = fields.Selection([('oui','Oui'),('non','Non')],"Retenu")
    adjudicataire_id = fields.Many2one('res.partner', u"Adjudicataire", index=True)
    montant          = fields.Float(u"Montant du lot"   , digits=(14,2))
    nomdulot         = fields.Char(u"Nom du lot")
    notes            = fields.Text(u"Notes")


class IsDossierNote(models.Model):
    _name = 'is.dossier.note'
    _description = "Dossier Note"
    _order = 'date'
    _rec_name = 'date'

    dossier_id = fields.Many2one('is.dossier', 'Dossier', required=True)
    date       = fields.Date("Date")
    auteur     = fields.Char("Auteur")
    titre      = fields.Char("Titre")
    texte_note = fields.Text("Note")


class IsDossierReference(models.Model):
    _name = 'is.dossier.reference'
    _description = u"Référence Dossier"
    _order = 'name'

    name        = fields.Char(u"Référence dossier", required=True, index=True)
    idreference = fields.Integer(u"idreference")


class IsDossierContratTraitance(models.Model):
    _name = 'is.dossier.contrat.traitance'
    _description = u"Traitance Contrat"
    _order = 'name'

    name   = fields.Char(u"Traitance", required=True, index=True)
    code   = fields.Char(u"Code")
    active = fields.Boolean(u"Active", default=True)


class IsDossierContrat(models.Model):
    _name = 'is.dossier.contrat'
    _description = u"Contrat"
    _order = 'name desc'


    @api.depends('detail_ids')
    def compute_restant_ht(self):
        for obj in self:
            restant_ht    = 0
            facturable_ht = 0
            if obj.signe=="oui":
                for line in obj.detail_ids:
                    if line.facturable=="oui":
                        restant_ht += line.montant_ht-line.montant_ht*line.facture_recursif/100.0
                        facturable_ht += line.montant_ht*(line.a_facturer-line.facture_recursif)/100.0
            obj.restant_ht    = restant_ht
            obj.facturable_ht = facturable_ht
        return True


    @api.depends('detail_ids')
    def compute_numligne_max(self):
        for obj in self:
            numligne_max=0
            for line in obj.detail_ids:
                if line.numligne>numligne_max:
                    numligne_max=line.numligne
            obj.numligne_max=numligne_max


    @api.multi
    def _get_numcontrat(self):
        context = self._context
        numcontrat=''
        if 'active_id' in context:
            dossier_id = context['active_id']
            dossiers=self.env['is.dossier'].search([('id','=',dossier_id)])
            for dossier in dossiers:
                numaff = dossier.numaff
                prefix = numaff.replace('A','CT')
                dossiers=self.env['is.dossier.contrat'].search([('dossier_id','=',dossier_id)])
                nb = len(dossiers) + 1
                numcontrat = prefix+'-'+str(nb)
        return numcontrat


    name          = fields.Char(u"Numero de contrat", required=True, index=True, default=lambda self: self._get_numcontrat())
    signe         = fields.Selection([('oui','Oui'),('non','Non')],"Signé")
    dossier_id    = fields.Many2one('is.dossier', u"Dossier", index=True, required=True, ondelete='cascade')
    client_id     = fields.Many2one('res.partner', u"Client", index=True)
    description   = fields.Char(u"Description du contrat")
    notes         = fields.Text(u"Notes")
    condreglement = fields.Text(u"Condition de règlement")
    delaipaiement = fields.Text(u"Délai de paiement")
    date          = fields.Date(u"Date de signature")
    refclient     = fields.Char(u"N°Commande client", help=u"Utilisé pour les marchés publiques")
    num_marche    = fields.Char(u"N°Marché"    , help=u"Utilisé pour les marchés publiques")
    #bon_commande  = fields.Char(u"N°Commande"  , help=u"Utilisé pour les marchés publiques")
    code_service  = fields.Char(u"Code service", help=u"Utilisé pour les marchés publiques")
    traitance_id  = fields.Many2one('is.dossier.contrat.traitance', u"Traitance", index=True)
    invoice_ids   = fields.One2many('account.invoice', 'is_contrat_id', u'Factures', domain=[('state','not in',['cancel'])], readonly=True)
    nb_factures   = fields.Integer(u"Nombre de factures")
    detail_ids    = fields.One2many('is.dossier.contrat.detail', 'contrat_id', u'Détail')
    restant_ht    = fields.Float(u"Restant HT"   , digits=(14,2), compute='compute_restant_ht', readonly=True, store=True)
    facturable_ht = fields.Float(u"Facturable HT", digits=(14,2), compute='compute_restant_ht', readonly=True, store=True)
    heure_ids     = fields.One2many('is.salarie.heure', 'contrat_id', u'Heures', readonly=True)
    tva_id        = fields.Many2one('is.tva', u'TVA')
    anomalie      = fields.Float(u"Anomalie", default=0)
    state         = fields.Selection([('en_cours','En cours'),('termine','Terminé')],"Etat", default='en_cours')

    numligne_max  = fields.Integer(u"N° Ligne maxi", compute='compute_numligne_max')



    # def get_invoice_line_recursif(self, compteur, niveau, lines, sens, invoice, invoices):
    #     for obj in self:
    #         for line in invoice.invoice_line_ids:
    #             if line.is_invoice_id:
    #                 invoices.append(line.is_invoice_id.name)
    #             txt_phase = line.is_contrat_detail_id.phase_id.txt_phase
    #             quantity  = line.quantity

    #             line_sens=sens
    #             if line.price_unit<0:
    #                 line_sens=-line_sens
    #             if invoice.type=="out_refund":
    #                 line_sens=-line_sens
    #             vals={
    #                 "niveau"      : niveau,
    #                 "invoice_name": invoice.name,
    #                 "invoice_type": invoice.type,
    #                 "is_contrat_detail_id": line.is_contrat_detail_id,
    #                 "txt_phase"   : txt_phase,
    #                 "price_unit"  : line.price_unit,
    #                 "sens"        : line_sens,
    #                 "quantity"    : quantity,
    #             }
    #             lines.append(vals)
    #             if line.is_invoice_id and line.is_invoice_id.state not in ['draft','cancel']:
    #                 res = obj.get_invoice_line_recursif(compteur+1, niveau+1, lines, line_sens, line.is_invoice_id, invoices)
    #                 lines    = res.get("lines"   , [])
    #                 compteur = res.get("compteur", 0)
    #                 invoices = res.get("invoices", [])
    #         res={
    #             "lines"   : lines,
    #             "compteur": compteur,
    #             "invoices": invoices,
    #         }
    #         return res


    # def compute_facture_recursif(self):
    #     for obj in self:
    #         lines=[]
    #         compteur=1
    #         for invoice in obj.invoice_ids:
    #             invoices=[]
    #             res = obj.get_invoice_line_recursif(compteur, 1, lines, 1, invoice, invoices)
    #             lines    = res.get("lines"   , [])
    #             compteur = res.get("compteur", 0)
    #             invoices = res.get("invoices", [])
    #         res={}
    #         for line in lines:
    #             is_contrat_detail_id = line["is_contrat_detail_id"]
    #             if is_contrat_detail_id:
    #                 if is_contrat_detail_id not in res:
    #                     res[is_contrat_detail_id]=0
    #                 res[is_contrat_detail_id]+=line["sens"]*line["quantity"]


    #         for line in obj.detail_ids:
    #             if line in res:compute_restant_ht(self)
# TODO : La numérotation automatique pose problème à l'importation et pour faire le lien avec les factures si modification de celle-ci
#    @api.multi
#    def write(self, vals):
#        res=super(IsDossierContrat, self).write(vals)
#        for obj in self:
#            numligne=10line.montant_ht*(line.a_facturer-line.facture_recursif)/100.0
#                line.numligne = numligne
#                numligne+=10
#        return res


    @api.multi
    def update_restant_ht_action(self):
        for obj in self:
            obj.compute_restant_ht()
        # for obj in self:

            # if len(obj.invoice_ids)>13:
            #     ct+=1

            # if len(obj.invoice_ids)==19:
            #     dt = datetime.now()
            #     obj.compute_restant_ht()
            #     tps = (datetime.now()-dt).total_seconds()
            # ct+=1

    @api.multi
    def update_reellement_facture_action(self):
        for obj in self:
            obj.update_reellement_facture_contrat(update_facture=True)
            obj.compute_restant_ht()

    def update_reellement_facture_contrat(self, update_facture=False):
        """
        Pour chaque contrat:
        1. Update le nombre de factures dans le contrat
        2. Update le reellement facturé de toutes les factures du contrat (OPTIONNEL)
        3. Update le pourcentage de réellement facturé pour chaque service dans ce contrat
        """
        t_start = datetime.now()
        for obj in self:
            anomalie = 0

            # 1. mettre à jour le nombre de factures dans ce contrat
            obj.nb_factures = len(obj.invoice_ids)

            # 2. met à jour le reellement facturé de toutes les factures
            # (seulement dans le cas de debug où on clique sur "actualiser reellement facture")
            if update_facture:
                obj.invoice_ids.update_deja_facture_action()

            # 3. Pour chaque service, récupérer le pourcentage de déjà facturé dans ce contrat
            # Pour cela: 
            #  3.1. récupérer tous les services et les factures
            #  3.2. mettre à jour le pourcentage de déjà facturé en parcourant chaque service de chaque facture
            factures_recursif = {} 

            # --- 3.1. récupérer tous les services et les factures --- 
            factures = [f for f in obj.invoice_ids if f.state not in ['draft','cancel']]
            services = obj.detail_ids
            # Initialisation du dict de % de déjà facturé
            factures_recursif = { line : 0 for line in services } 
            
             # --- 3.2. maj du % de déjà facturé en parcourant chaque service de chaque facture --- 
            for f in factures:
                # Mettre à jour le reellement facturé pour les services de cette facture
                for line in f.invoice_line_ids:
                    if line.is_contrat_detail_id:
                        # Somme des réellement facturé de chaque facture * le prix
                        signe_line = 1
                        if line.price_unit < 0:
                            signe_line = 1    
                        factures_recursif[line.is_contrat_detail_id] += signe_line*line.is_reellement_facture

            # Le transformer en pourcentage
            for s in services:
                s.facture_recursif = factures_recursif[s]*100
                anomalie += abs(s.facture_recursif - s.a_facturer)*s.montant_ht/100
                    
            obj.anomalie = anomalie
        t_end = datetime.now()
        #print("Mise à jour de tous les contrats en %.6f secondes" %((t_end - t_start).total_seconds()))

    @api.multi
    def acceder_contrat_action(self, vals):
        for obj in self:
            res= {
                'name': 'Contrat',
                'view_mode': 'form',
                'view_type': 'form',
                'res_model': 'is.dossier.contrat',
                'res_id': obj.id,
                'type': 'ir.actions.act_window',
            }
            return res


    @api.multi
    def generer_facture_action(self, vals):
        for obj in self:
            # mettre à jour le nombre de factures dans ce contrat
            obj.nb_factures = len(obj.invoice_ids)


            # #** Recherche si il y a des lignes a facturer **********************
            # test=False
            # for line in obj.detail_ids:
            #     if line.a_facturer>line.facture:
            #         test=True
            # if test==False:
            #     raise Warning(u"Aucune ligne à facturer sur ce contrat")
            # #*******************************************************************


            #** Recherche du numéro de facture *********************************
            prefix = _get_prefix_annee()+'FC'
            filtre=[
                ('number','like',prefix),
            ]
            invoices = self.env['account.invoice'].search(filtre,limit=1,order='number desc')
            suffix=1
            for invoice in invoices:
                suffix = int(invoice.number[-3:])+1
            suffix = ('000'+str(suffix))[-3:]
            number  = prefix+suffix
            #*******************************************************************


            #** Création entête facture ****************************************
            invoice_type="out_invoice"
            sens_facture=1
            if obj.facturable_ht<0:
                invoice_type="out_refund"
                sens_facture=-1
            vals={
                'name'              : number,
                'origin'            : obj.name,
                'move_name'         : number,
                'number'            : number,
                'partner_id'        : obj.client_id.id,
                'is_contrat_id'     : obj.id,
                'fiscal_position_id': 1,
                'is_tva_id'         : obj.tva_id.id,
                'type'              : invoice_type,
                'state'             : 'draft',
            }

            invoice=self.env['account.invoice'].create(vals)
            invoice_id = invoice.id
            #*******************************************************************

            #** Création lignes du contrat *************************************
            for line in obj.detail_ids:
                if line.commentaire:
                    vals={
                        'invoice_id': invoice_id,
                        'is_contrat_detail_id': line.id,
                        'display_type': 'line_note',
                        'name'      : line.texte,
                        'quantity'  : False,
                        # 'price_unit': False,
                        'account_id': False,
                    }
                    invoice_line=self.env['account.invoice.line'].create(vals)
                else:
                    quantity = line.a_facturer/100.0
                    vals={
                        'invoice_id': invoice_id,
                        'is_contrat_detail_id': line.id,
                        'product_id': 1,
                        'name'      : line.texte,
                        'quantity'  : quantity,
                        'price_unit': False,
                        'account_id': 622, #701100
                    }
                    invoice_line=self.env['account.invoice.line'].create(vals)
                    line_res = invoice_line.uptate_onchange_product_id()
                    vals={
                        'name'      : line.texte,
                        #'price_unit': sens_facture*line.montant_ht,
                        'price_unit': sens_facture*line.montant_ht,
                    }
                    invoice_line.write(vals)
            #*******************************************************************

            #** Ajout des factures réalisées ***********************************
            filtre=[
                ('is_contrat_id','=',obj.id),
                ('date_invoice','<=',date.today()),
                ('state','not in',['cancel']),
                ('id','!=',invoice_id),
            ]
            factures = self.env['account.invoice'].search(filtre,order='date_invoice')
            for facture in factures:
                sens_ligne_facture = 1
                if facture.type == "out_refund":
                    sens_ligne_facture = -1
                vals={
                    'invoice_id'   : invoice_id,
                    'product_id'   : 2,
                    'name'         : facture.number,
                    'quantity'     : 1,
                    #'price_unit'   : -facture.amount_untaxed,
                    'price_unit'   : 0,
                    'account_id'   : 622, #701100
                    'is_invoice_id': facture.id,
                }
                invoice_line=self.env['account.invoice.line'].create(vals)
                line_res = invoice_line.uptate_onchange_product_id()
                vals={
                    'name'      : facture.number,
                    #'price_unit': -facture.amount_untaxed,
                    # On veut que le montant final d'une invoice soit positif quelque soit le type d'invoice
                    # Donc on veut que les lignes des invoices précédentes soient en moins si sont type est different
                    # du type de l'invoice donc "-(sens_facture*sens_ligne_facture)"
                    'price_unit'   : -(sens_facture*sens_ligne_facture)*facture.is_montant_hors_revision, # Ceci donne tout bon! 3/3
                }
                invoice_line.write(vals)
            #*******************************************************************

            #** Recalcul de la TVA et validation de la facture *****************
            res_validate = invoice.update_tva_account_action()
            res_validate = invoice.compute_taxes()
            try:
               res_validate = invoice.action_invoice_open()
            except:
               continue
            #*******************************************************************

            #*******************************************************************
            invoice.update_deja_facture_action()
            obj.update_reellement_facture_contrat()
            obj.compute_restant_ht()


class IsDossierContratDetail(models.Model):
    _name = 'is.dossier.contrat.detail'
    _description = u"Détail Contrat"
    _order = 'contrat_id,numligne,id'
    _rec_name = 'numligne'


    @api.depends('montant1','montant2','montant3','montant4','avancement','a_facturer','facture_recursif')
    def _compute_montant_ht(self):
        for obj in self:
            obj.montant_ht = obj.montant1 + obj.montant2 + obj.montant3 + obj.montant4
            obj.reste_facturable = obj.avancement - obj.facture_recursif


    @api.depends('a_facturer')
    def _compute_facture(self):
        for obj in self:
            obj.montant_a_facturer = obj.montant_ht*(obj.a_facturer-obj.facture_recursif)/100
            obj.montant_facture    = obj.montant_ht*(obj.facture_recursif)/100
            encours=0
            montant_encours=0
            restant=0
            montant_restant=0
            if obj.facturable=='oui':
                if obj.avancement==100:
                    restant=100-obj.facture_recursif
                    montant_restant = restant*obj.montant_ht/100
                else:
                    encours=obj.avancement-obj.facture_recursif
                    montant_encours = encours*obj.montant_ht/100
            obj.encours = encours
            obj.montant_encours = montant_encours
            obj.restant = restant
            obj.montant_restant = montant_restant


    contrat_id  = fields.Many2one('is.dossier.contrat', 'Contrat', required=True, ondelete='cascade')
    state       = fields.Selection(related='contrat_id.state')
    dossier_id  = fields.Many2one(related='contrat_id.dossier_id')
    description = fields.Char(related='contrat_id.description')

    numligne    = fields.Integer(u"Ligne", default=0) #, default=lambda self: self._get_numligne(self.contrat_id.numligne_max))
    commentaire = fields.Boolean(u"Commentaire", default=False)
    phase_id    = fields.Many2one('is.dossier.contrat.phase', 'Phase')
    det_phase   = fields.Char(u"Description Phase", related='phase_id.det_phase')

    texte       = fields.Char(u"Description")
    dateremise  = fields.Date(u"Date")
    facture_le  = fields.Date(u"Pour le")

    stxt1       = fields.Char(u"Description 1")
    montant1    = fields.Float(u"Montant 1", digits=(14,4))
    codass1_id  = fields.Many2one('is.dossier.code.assurance', 'Code assurance 1')

    stxt2       = fields.Char(u"Description 2")
    montant2    = fields.Float(u"Montant 2", digits=(14,4))
    codass2_id  = fields.Many2one('is.dossier.code.assurance', 'Code assurance 2')

    stxt3       = fields.Char(u"Description 3")
    montant3    = fields.Float(u"Montant 3", digits=(14,4))
    codass3_id  = fields.Many2one('is.dossier.code.assurance', 'Code assurance 3')

    stxt4       = fields.Char(u"Description 4")
    montant4    = fields.Float(u"Montant 4", digits=(14,4))
    codass4_id  = fields.Many2one('is.dossier.code.assurance', 'Code assurance 4')

    montant_ht  = fields.Float(u"Montant HT", digits=(14,2), compute='_compute_montant_ht', readonly=True, store=True)

    avancement  = fields.Float(u"% avancement", digits=(14,4))
    nota        = fields.Char(u"Nota")
    facturable  = fields.Selection([('oui','Oui'),('non','Non')],"Facturable", default='oui')
    a_facturer  = fields.Float(u"% à facturer", digits=(14,4))
    facture            = fields.Float(u"% facturé"         , compute='_compute_facture', readonly=True, store=False, digits=(14,4))
    montant_a_facturer = fields.Float(u"Montant à facturer", compute='_compute_facture', readonly=True, store=False)
    montant_facture    = fields.Float(u"Montant facturé"   , compute='_compute_facture', readonly=True, store=False)

    facture_recursif   = fields.Float("% facturé récursif", digits=(14,4))

    reste_facturable   = fields.Float(u"% Reste Facturable", digits=(14,4), compute='_compute_montant_ht', readonly=True, store=True)

    encours                    = fields.Float(u"% En cours"                , digits=(14,4), compute='_compute_facture', readonly=True, store=False)
    montant_encours            = fields.Float(u"Montant en cours"          , digits=(14,2), compute='_compute_facture', readonly=True, store=False)
    restant                    = fields.Float(u"% Restant à facturer"      , digits=(14,4), compute='_compute_facture', readonly=True, store=False)
    montant_restant            = fields.Float(u"Montant restant à facturer", digits=(14,2), compute='_compute_facture', readonly=True, store=False)


    @api.model
    def create(self, vals):
        res=super(IsDossierContratDetail, self).create(vals)
        res.numligne = res.contrat_id.numligne_max+10
        return res


    @api.multi
    def unlink(self):
        for obj in self:
            lines = self.env['account.invoice.line'].search([('is_contrat_detail_id','=',obj.id)])
            if len(lines)>0:
                factures=[]
                for line in lines:
                    factures.append(line.invoice_id.move_name)
                factures=", ".join(factures)
                raise Warning("Il n'est pas possible de supprimer la ligne '%s' car elle est utilsée dans les factures %s"%(obj.texte,factures))
        return super(IsDossierContratDetail, self).unlink()


class IsDossierContratPhase(models.Model):
    _name = 'is.dossier.contrat.phase'
    _description = u"Phase Contrat"
    _order = 'txt_phase'
    _rec_name = 'txt_phase'

    ref_phase       = fields.Integer(u"Ref"     , required=True, index=True)
    txt_phase       = fields.Char(u"Code"       , required=True, index=True)
    det_phase       = fields.Char(u"Description", required=True, index=True)


class IsDossierCodeAssurance(models.Model):
    _name = 'is.dossier.code.assurance'
    _description = u"Code Assurance"
    _order = 'ordre,code'
    _rec_name = 'code'

    ordre       = fields.Char(u"Ordre"      , required=True, index=True, default='x')
    code        = fields.Char(u"Code"       , required=True, index=True)
    description = fields.Char(u"Description", required=True, index=True)


class IsSalarieHeure(models.Model):
    _name = 'is.salarie.heure'
    _description = u"Heures des salariés"
    _order = 'date_travaillee desc'
    _rec_name = 'date_travaillee'

    date_travaillee = fields.Date(u"Date travaillée", required=True, index=True, default=lambda *a: fields.Date.today())
    user_id         = fields.Many2one('res.users', 'Utilisateur', required=True, default=lambda self: self.env.user)
    login           = fields.Char(u"Login")
    contrat_id      = fields.Many2one('is.dossier.contrat', 'Contrat', required=True, ondelete='cascade')
    nb_heures       = fields.Float(u"Nb heures")
    note            = fields.Char(u"Note")
    id_heures       = fields.Integer(u"id_heures", index=True)


class IsDeclarationMAF(models.Model):
    _name = 'is.declaration.maf'
    _description = u"Déclaration MAF"

    name          = fields.Char(u"Année", required=True, index=True)
    type_document = fields.Selection([('maf',u'Déclaration MAF'),('dget','DGET')],"Type de document", default='maf')

    @api.multi
    def getCodeAssurance(self,code_id):
        cr = self._cr
        SQL="SELECT code,ordre FROM is_dossier_code_assurance where id="+str(code_id)
        cr.execute(SQL)
        rows = cr.fetchall()
        code='0'
        ordre=''
        for row in rows:
            code  = row[0]
            ordre = row[1]
        return code, ordre


    @api.multi
    def get_contrats(self,obj):
        cr = self._cr
        SQL="""
            select 
                idcd.codass1_id,
                idcd.codass2_id,
                idcd.codass3_id,
                idcd.codass4_id,
                idcd.montant1,
                idcd.montant2,
                idcd.montant3,
                idcd.montant4,
                ai.name,
                ail.sequence,
                idc.name,
                id.numaff,
                rp.name,
                ail.is_contrat_detail_id,
                ail.price_subtotal,
                (ail.quantity-ail.is_deja_facture),
                ail.is_invoice_id,

                ai.is_montant_revision,

                (ail.quantity-ail.is_deja_facture)*idcd.montant1/NullIf(((ail.quantity-ail.is_deja_facture)*ail.price_unit),0) rapport1,
                (ail.quantity-ail.is_deja_facture)*idcd.montant2/NullIf(((ail.quantity-ail.is_deja_facture)*ail.price_unit),0) rapport2,
                (ail.quantity-ail.is_deja_facture)*idcd.montant3/NullIf(((ail.quantity-ail.is_deja_facture)*ail.price_unit),0) rapport3,
                (ail.quantity-ail.is_deja_facture)*idcd.montant4/NullIf(((ail.quantity-ail.is_deja_facture)*ail.price_unit),0) rapport4,

                (ail.quantity-ail.is_deja_facture)*ail.price_unit/NullIf(ai.is_montant_hors_revision,0) rapport_total,

                (ail.quantity-ail.is_deja_facture)*idcd.montant1 + ((ail.quantity-ail.is_deja_facture)*idcd.montant1/NullIf(((ail.quantity-ail.is_deja_facture)*ail.price_unit),0)) * ((ail.quantity-ail.is_deja_facture)*ail.price_unit/NullIf(ai.is_montant_hors_revision,0)) * ai.is_montant_revision revision1,
                (ail.quantity-ail.is_deja_facture)*idcd.montant2 + ((ail.quantity-ail.is_deja_facture)*idcd.montant2/NullIf(((ail.quantity-ail.is_deja_facture)*ail.price_unit),0)) * ((ail.quantity-ail.is_deja_facture)*ail.price_unit/NullIf(ai.is_montant_hors_revision,0)) * ai.is_montant_revision revision2,
                (ail.quantity-ail.is_deja_facture)*idcd.montant3 + ((ail.quantity-ail.is_deja_facture)*idcd.montant3/NullIf(((ail.quantity-ail.is_deja_facture)*ail.price_unit),0)) * ((ail.quantity-ail.is_deja_facture)*ail.price_unit/NullIf(ai.is_montant_hors_revision,0)) * ai.is_montant_revision revision3,
                (ail.quantity-ail.is_deja_facture)*idcd.montant4 + ((ail.quantity-ail.is_deja_facture)*idcd.montant4/NullIf(((ail.quantity-ail.is_deja_facture)*ail.price_unit),0)) * ((ail.quantity-ail.is_deja_facture)*ail.price_unit/NullIf(ai.is_montant_hors_revision,0)) * ai.is_montant_revision revision4





            from account_invoice_line ail inner join account_invoice             ai on ail.invoice_id=ai.id
                                          inner join is_dossier_contrat         idc on ai.is_contrat_id=idc.id
                                          inner join is_dossier_contrat_detail idcd on ail.is_contrat_detail_id=idcd.id
                                          inner join res_partner                 rp on ai.partner_id=rp.id
                                          inner join is_dossier                  id on idc.dossier_id=id.id
            where 
                ai.date_invoice>='"""+str(obj.name)+"""-01-01' and
                ai.date_invoice<='"""+str(obj.name)+"""-12-31' and
                ai.state not in ('draft','cancel')
                -- and id.numaff='32A011'
            order by ai.name,ail.sequence
        """
        # and idc.name in ('34CT015-6','34CT018-1')
        cr.execute(SQL)
        rows = cr.fetchall()
        nb=len(rows)
        r={}
        for row in rows:
            contrat = row[10]
            if contrat not in r:
                r[contrat]={}
            qty = row[15]
            if qty:
                if row[0]:
                    if row[0] not in r[contrat]:
                        r[contrat][row[0]]=0
                    r[contrat][row[0]]+=row[23] or 0

                if row[1]:
                    if row[1] not in r[contrat]:
                        r[contrat][row[1]]=0
                    r[contrat][row[1]]+=row[24] or 0

                if row[2]:
                    if row[2] not in r[contrat]:
                        r[contrat][row[2]]=0
                    r[contrat][row[2]]+=row[25] or 0

                if row[3]:
                    if row[3] not in r[contrat]:
                        r[contrat][row[3]]=0
                    r[contrat][row[3]]+=row[26] or 0
                if not row[0] and not row[1] and not row[2] and not row[3]:
                    if 0 not in r[contrat]:
                        r[contrat][0]=0
                    r[contrat][0]+=row[17] or 0
        keys=sorted(r.keys(), reverse=True)
        total=0
        res=[]
        recap={}
        for k in keys:
            lines={}
            for code_id in r[k]:
                montant = r[k][code_id]
                code,ordre = self.getCodeAssurance(code_id)
                if code!="ND":
                    lines[ordre]=[code,montant]
            keys2=sorted(lines.keys(), reverse=False)
            for k2 in keys2:
                line    = lines[k2]
                code    = line[0].strip()
                montant = line[1]
                contrats = self.env['is.dossier.contrat'].search([('name','=',k)],order="name",limit=1)
                #print('TEST 3',line,code,montant,contrats)
                if contrats:
                    contrat=contrats[0]
                    traitance = contrat.traitance_id.code or ''
                    cle=code+' - '+traitance
                    if cle not in recap:
                        recap[cle]=0
                    recap[cle]+=montant
                    if montant!=0:
                        vals={
                            'contrat'  : contrat,
                            'code'     : code,
                            'traitance': traitance,
                            'montant'  : montant,
                        }
                        res.append(vals)
                    total+=r[k][code_id]
        keys=sorted(recap.keys(), reverse=False)
        recap2=[]
        for k in keys:
            recap2.append([k,recap[k]])
        return [res,recap2]

    @api.multi
    def export_excel_action(self):
        for obj in self:
            res = obj.get_contrats(obj)

            excel_attachment_id = False
            #** Création du workbook ******************************************
            name  = 'declaration-maf-%s.xlsx'%obj.name
            path = '/tmp/%s'%name
            workbook = Workbook()
            ws = workbook.active
            ws.title = "Déclaration MAF"
            ws.sheet_view.showGridLines = False
            border1 = borders.Side(style = None, color = 'FF000000', border_style = 'thin')
            thin_border = Border(left = border1, right = border1, bottom = border1, top = border1)
            #******************************************************************

            #** Ligne d'entête ************************************************
            titres=["N°affaire","N°contrat", "Nom affaire", "Code MAF", "Contrat HT", "Affaire HT"]
            column = 1
            row = 1
            for titre in titres:
                clr_background = PatternFill(start_color='FFFACD', end_color='FFFACD', fill_type="solid")
                cell = ws.cell(row=row, column=column, value=titre)
                cell.fill = clr_background
                cell.alignment = Alignment(horizontal='center')
                cell.font = Font(name='Calibri', bold=True)
                cell.border = thin_border
                column+=1
            #******************************************************************

            #** Largeur des colonnes ******************************************
            ws.column_dimensions['A'].width=12
            ws.column_dimensions['B'].width=15
            ws.column_dimensions['C'].width=40
            ws.column_dimensions['D'].width=12
            ws.column_dimensions['E'].width=15
            ws.column_dimensions['F'].width=15
            #******************************************************************


            #** Totaux des affaires *******************************************
            lines = res[0]
            totaux={}
            for line in lines:
                numaff = line["contrat"].dossier_id.numaff
                if numaff not in totaux:
                    totaux[numaff] = 0
                totaux[numaff]+=line["montant"]
            #******************************************************************


            #** Lignes ********************************************************
            color1 = PatternFill(start_color='DEEBF7', end_color='DEEBF7', fill_type="solid")
            color2 = PatternFill(start_color='D0CECE', end_color='D0CECE', fill_type="solid")
            row+=1
            color=color1
            mem = ""
            first=True
            for line in lines:
                contrat = line["contrat"]
                if mem!=contrat.dossier_id.numaff:
                    first=True
                    total = 0
                    mem = contrat.dossier_id.numaff
                    if color==color1:
                        color=color2
                    else:
                        color=color1

                cell = ws.cell(row=row, column=1, value=contrat.dossier_id.numaff)
                cell.alignment = Alignment(horizontal='center')
                cell.font = Font(name='Calibri', bold=False)
                cell.border = thin_border
                cell.fill = color

                cell = ws.cell(row=row, column=2, value=contrat.name)
                cell.alignment = Alignment(horizontal='center')
                cell.font = Font(name='Calibri', bold=False)
                cell.border = thin_border
                cell.fill = color

                cell = ws.cell(row=row, column=3, value=contrat.dossier_id.nom)
                cell.alignment = Alignment(horizontal='left')
                cell.font = Font(name='Calibri', bold=False)
                cell.border = thin_border
                cell.fill = color

                cell = ws.cell(row=row, column=4, value=line["code"])
                cell.alignment = Alignment(horizontal='center')
                cell.font = Font(name='Calibri', bold=False)
                cell.border = thin_border
                cell.fill = color

                cell = ws.cell(row=row, column=5, value=line["montant"])
                cell.alignment = Alignment(horizontal='right')
                cell.font = Font(name='Calibri', bold=False)
                cell.border = thin_border
                cell.fill = color
                cell.number_format = '# ##0.00'  # Number formatting


                if first:
                    value = totaux[contrat.dossier_id.numaff]
                    cell = ws.cell(row=row, column=6, value=value)
                    cell.alignment = Alignment(horizontal='right')
                    cell.font = Font(name='Calibri', bold=False)
                    cell.border = thin_border
                    cell.fill = color
                    cell.number_format = '# ##0.00'  # Number formatting

                first=False
                row+=1
            #******************************************************************

            #** Récapitulatif *************************************************
            row+=2
            titres=["Code MAF", "Montant HT"]
            column = 4
            for titre in titres:
                clr_background = PatternFill(start_color='FFFACD', end_color='FFFACD', fill_type="solid")
                cell = ws.cell(row=row, column=column, value=titre)
                cell.fill = clr_background
                cell.alignment = Alignment(horizontal='center')
                cell.font = Font(name='Calibri', bold=True)
                cell.border = thin_border
                column+=1
            row+=1
            recap2 = res[1]
            cell_start = cell_end = False
            for line in recap2:
                cell = ws.cell(row=row, column=4, value=line[0])
                cell.alignment = Alignment(horizontal='left')
                cell.font = Font(name='Calibri', bold=False)
                cell.border = thin_border

                cell = ws.cell(row=row, column=5, value=line[1])
                cell.alignment = Alignment(horizontal='right')
                cell.font = Font(name='Calibri', bold=False)
                cell.border = thin_border
                cell.number_format = '# ##0.00'  # Number formatting
                #** cell de début et de fin pour faire la somme ***************
                if cell_start==False:
                    cell_start = cell
                cell_end = cell
                row+=1
            #******************************************************************

            #** Total du récapitulatif ****************************************
            cell = ws.cell(row=row, column=4, value="Total : ")
            cell.alignment = Alignment(horizontal='left')
            cell.font = Font(name='Calibri', bold=True)
            cell.border = thin_border

            formula = "=SUM(%s:%s)"%(cell_start.coordinate, cell_end.coordinate)
            cell = ws.cell(row=row, column=5, value=formula)
            cell.alignment = Alignment(horizontal='right')
            cell.font = Font(name='Calibri', bold=True)
            cell.border = thin_border
            cell.number_format = '# ##0.00'  # Number formatting
            #******************************************************************

            #** Mise en page **************************************************
            #ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
            ws.sheet_properties.pageSetUpPr.fitToPage = True
            ws.page_setup.fitToHeight = False
            ws.oddFooter.center.text = "Page &[Page] of &N"
            ws.page_margins = PageMargins(left=0.3, right=0.3, top=0.3, bottom=0.7, header=0, footer=0.3)
            ws.print_title_rows = '1:1' # the first two rows 
            ws.print_options.horizontalCentered = True
            #******************************************************************

            #** Fixer les lignes et les colonnes ******************************
            cel = ws['A2']
            ws.freeze_panes = cel
            #******************************************************************

            # #** Enregistrement du workbook ************************************
            workbook.save(path)
            # #******************************************************************

            # ** Creation ou modification de la pièce jointe ******************
            user  = self.env['res.users'].browse(self._uid)
            attachment_obj = self.env['ir.attachment']
            attachments = attachment_obj.search([('res_id','=',user.id),('name','=',name)])
            xlsx = open(path,'rb').read()
            vals = {
                'name':        name,
                'type':        'binary',
                'res_id':      user.id,
                'datas':       base64.b64encode(xlsx),
            }
            attachment_id=False
            if attachments:
                for attachment in attachments:
                    attachment.write(vals)
                    attachment_id=attachment.id
            else:
                attachment = attachment_obj.create(vals)
                attachment_id=attachment.id
            #*******************************************************************

            ##** Envoi de l'action dans le navigateur **************************
            if attachment_id:
                vals={
                    "type": 'ir.actions.act_url',
                    "url" : '/web/content/%s?download=true'%attachment_id,
                }
                return vals

    # ExcelClick(ev) {
    #     const excel_attachment_id = ev.target.attributes.excel_attachment_id.value;
    #     this.action.doAction);
    # }







class IsConcours(models.Model):
    _name = 'is.concours'
    _description = u"Concours"
    _order = 'numconcours desc'
    _rec_name = 'numconcours'

    @api.multi
    def _get_numconcours(self):
        prefix = _get_prefix_annee()+'CO'
        rows=self.env['is.concours'].search([('numconcours','like',prefix)],order="numconcours desc",limit=1)
        numconcours=1
        for row in rows:
            numconcours = int(row.numconcours[-3:])
            numconcours+=1
        numconcours=prefix+('000'+str(numconcours))[-3:]
        return numconcours

    numconcours    = fields.Char(u"Numero de concours", required=True, index=True, default=lambda self: self._get_numconcours())
    organisme_id   = fields.Many2one('res.partner', u"Organisme", domain=[('is_company','=',True)])
    nom            = fields.Char(u"Nom du concours")
    libelle_exacte = fields.Text(u"Libelle exacte")
    lieu           = fields.Char(u"Lieu")
    dateparution   = fields.Date(u"Date de parution")
    daterendu      = fields.Date(u"Date rendu")
    journal        = fields.Char(u"Journal")
    lienannonce    = fields.Char(u"Lien annonce")
    qui_ids        = fields.One2many('is.concours.qui', 'concours_id', u'Equipes')


class IsConcoursQui(models.Model):
    _name = 'is.concours.qui'
    _description = u"Concours Qui"
    _order = 'nom'
    _rec_name = 'nom'

    concours_id    = fields.Many2one('is.concours', 'Concours', required=True, ondelete='cascade')
    nom            = fields.Char(u"Nom", required=True, index=True)
    dateenvoi      = fields.Date(u"Date d'envoi")
    retenu         = fields.Selection([('oui','Oui'),('non','Non')],"Gagné" , default='non')
    gagne          = fields.Selection([('oui','Oui'),('non','Non')],"Retenu", default='non')
    equipes        = fields.Text(u"Equipes")
    idconcours_qui = fields.Integer(u"idconcours_qui", index=True)


    


